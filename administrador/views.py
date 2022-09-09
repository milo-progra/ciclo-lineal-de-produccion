

from ast import Return
from datetime import datetime
from pipes import Template
from urllib import request
from django.shortcuts import render
from app.models import RegistroTrabajador, Etapa, Entrada, Salida, Oportunidades, Empresa, AreaEmpresa
from django.db.models import Count
import collections
import numpy as np
from Levenshtein import distance, editops, apply_edit, jaro
from django.views.generic import TemplateView
from openpyxl import Workbook
from django.http.response import HttpResponse
from .models import LogTelegram





# Create your views here.

def homeAdmin(request):
    registros = RegistroTrabajador.objects.filter(usuario=request.user)
    empresas = Empresa.objects.all()

    data = {

            
            'registros': registros,
            'empresas': empresas
            

    }
    return render(request,'home_admin.html', data)


def home_empresa(request, id):
    registros = RegistroTrabajador.objects.filter(usuario=request.user)
    empresas = Empresa.objects.all()
    empresa = Empresa.objects.filter(id_empresa = id)

    data = {
            'registros': registros,  
            'empresa':empresa,
            'empresas':empresas
    }

    return render(request,'empresa_1/home_empresa.html', data)   


def tablasExtraccion(request,id):
        if request.user.is_authenticated:
                empresas = Empresa.objects.all()
                empresa = Empresa.objects.filter(id_empresa = id)

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapas = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Extraccion materia prima")
                etapa = Etapa.objects.get(nombre = "Extraccion materia prima") #trar solo la ID de la etapa "Extraccion materia prima"
                #empresa = Empresa.objects.get(id)
                empresaArea = RegistroTrabajador.objects.all()
                area = AreaEmpresa.objects.filter(id_empresa = id)
                entradas = Entrada.objects.filter(etapa_id = etapa)


                #/////////////// Levenshtein ///////////////
                lista_u = []
                plabra_es = []
                lista_t = [] 
                nombre_espa = ""
                for e in entradas:
                        for a in area:
                                if e.id_area_id == a.id_area:        
                                        #print("las notas son: ",e.nombre) 
                                        e_nombre = e.nombre
                                        result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                                        #print(result)
                                        #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                                        if result > 1 :
                                                nombre_espa = e.nombre
                                                print(nombre_espa)
                                                separador = " "
                                                maximo_numero_de_separaciones = 2
                                                separado_por_espacios = nombre_espa.split(separador, maximo_numero_de_separaciones)
                                                plabra_es = plabra_es + separado_por_espacios
                                                
                                        else:
                                                lista_u.append(e_nombre)

                lista_t =  plabra_es + lista_u                       
                #print(lista_t)                    

                #////////////////////////////////////////
                
                result = (Entrada.objects
                .values('id_area')
                .annotate(dcount=Count('id_area'))
                .order_by()
                )
                #print("variable result!!!!!!!: ",result)

               

                theanswer = Entrada.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa) #requiere importar from django.db.models import Count
                #print(theanswer)
                salidas_count = Salida.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa)
                
                oportunidad_count = Oportunidades.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa)
              
                t = theanswer[0]
                                
                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea,
                'area' : area,
                't': t,
                'theanswer': theanswer,
                'salidas_count':salidas_count,
                'oportunidad_count':oportunidad_count,
                'empresas':empresas,
                'empresa':empresa,
                'lista_t':lista_t
                

                }
       
                return render(request,'empresa_1/tablas_extraccion.html', data)
        else:
                return render(request, 'empresa_1/tabla_extraccion.html')


def tablasDiseño(request,id):
        if request.user.is_authenticated:
                empresas = Empresa.objects.all()
                empresa = Empresa.objects.filter(id_empresa = id)

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapas = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Extraccion materia prima")
                etapa = Etapa.objects.get(nombre = "Diseño y produccion") #trar solo la ID de la etapa "Extraccion materia prima"
                #empresa = Empresa.objects.get(id)
                empresaArea = RegistroTrabajador.objects.all()
                
                area = AreaEmpresa.objects.filter(id_empresa = id)
                entradas = Entrada.objects.filter(etapa_id = etapa)
                
                
                result = (Entrada.objects
                .values('id_area')
                .annotate(dcount=Count('id_area'))
                .order_by()
                )
                print("variable result!!!!!!!: ",result)

               

                theanswer = Entrada.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa) #requiere importar from django.db.models import Count
                print(theanswer)
                salidas_count = Salida.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa)
                
                oportunidad_count = Oportunidades.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa)
              
                
                                
                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea,
                'area' : area,
                'theanswer': theanswer,
                'salidas_count':salidas_count,
                'oportunidad_count':oportunidad_count,
                'empresas':empresas,
                'empresa':empresa
                

                }
       
                return render(request,'empresa_1/tablas_diseño.html', data)
        else:
                return render(request, 'empresa_1/tabla_diseño.html')


def tablasLogistica(request,id):
        if request.user.is_authenticated:
                empresas = Empresa.objects.all()
                empresa = Empresa.objects.filter(id_empresa = id)
                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.get(nombre = "Logistica") #trar solo la ID de la etapa "Extraccion materia prima"
                empresaArea = RegistroTrabajador.objects.all()
                area = AreaEmpresa.objects.filter(id_empresa = id)
                theanswer = Entrada.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa) #requiere importar from django.db.models import Count
                salidas_count = Salida.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa)
                oportunidad_count = Oportunidades.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa)
                            
                data = {

                'registros': registros,
                'empresaArea':empresaArea,
                'area' : area,
                'theanswer': theanswer,
                'salidas_count':salidas_count,
                'oportunidad_count':oportunidad_count,
                'empresas':empresas,
                'empresa':empresa
        
                }
       
                return render(request,'empresa_1/tablas_logistica.html', data)
        else:
                return render(request, 'empresa_1/tabla_logistica.html')


def tablasCompra(request,id):
        if request.user.is_authenticated:
                empresas = Empresa.objects.all()
                empresa = Empresa.objects.filter(id_empresa = id)
                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.get(nombre = "Compra") #trar solo la ID de la etapa "Extraccion materia prima"
                empresaArea = RegistroTrabajador.objects.all()
                area = AreaEmpresa.objects.filter(id_empresa = id)
                theanswer = Entrada.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa) #requiere importar from django.db.models import Count
                salidas_count = Salida.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa)
                oportunidad_count = Oportunidades.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa)
                            
                data = {

                'registros': registros,
                'empresaArea':empresaArea,
                'area' : area,
                'theanswer': theanswer,
                'salidas_count':salidas_count,
                'oportunidad_count':oportunidad_count,
                'empresas':empresas,
                'empresa':empresa
        
                }
       
                return render(request,'empresa_1/tablas_compra.html', data)
        else:
                return render(request, 'empresa_1/tabla_compra.html')

def tablasUso(request,id):
        if request.user.is_authenticated:
                empresas = Empresa.objects.all()
                empresa = Empresa.objects.filter(id_empresa = id)
                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.get(nombre = "Uso consumo") #trar solo la ID de la etapa "Extraccion materia prima"
                empresaArea = RegistroTrabajador.objects.all()
                area = AreaEmpresa.objects.filter(id_empresa = id)
                theanswer = Entrada.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa) #requiere importar from django.db.models import Count
                salidas_count = Salida.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa)
                oportunidad_count = Oportunidades.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa)
                            
                data = {

                'registros': registros,
                'empresaArea':empresaArea,
                'area' : area,
                'theanswer': theanswer,
                'salidas_count':salidas_count,
                'oportunidad_count':oportunidad_count,
                'empresas':empresas,
                'empresa':empresa
        
                }
       
                return render(request,'empresa_1/tablas_uso.html', data)
        else:
                return render(request, 'empresa_1/tabla_uso.html')


def tablasFin(request,id):
        if request.user.is_authenticated:
                empresas = Empresa.objects.all()
                empresa = Empresa.objects.filter(id_empresa = id)
                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.get(nombre = "Fin de vida") #trar solo la ID de la etapa "Extraccion materia prima"
                empresaArea = RegistroTrabajador.objects.all()
                area = AreaEmpresa.objects.filter(id_empresa = id)
                theanswer = Entrada.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa) #requiere importar from django.db.models import Count
                salidas_count = Salida.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa)
                oportunidad_count = Oportunidades.objects.values('id_area').annotate(Count('id_area')).filter(etapa_id = etapa)
                            
                data = {

                'registros': registros,
                'empresaArea':empresaArea,
                'area' : area,
                'theanswer': theanswer,
                'salidas_count':salidas_count,
                'oportunidad_count':oportunidad_count,
                'empresas':empresas,
                'empresa':empresa
        
                }
       
                return render(request,'empresa_1/tablas_fin.html', data)
        else:
                return render(request, 'empresa_1/tabla_fin.html')



def promedioHome(request,id):
        if request.user.is_authenticated:
                empresa = Empresa.objects.filter(id_empresa = id)
                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.get(nombre = "Extraccion materia prima") #trar solo la ID de la etapa "Extraccion materia prima"
                area = AreaEmpresa.objects.filter(id_empresa = id)
                            
                data = {
                'registros': registros,
                'etapa':etapa,
                'area' : area,
                'empresa':empresa,    

                }
       
                return render(request,"empresa_1/promedios/promedio_home.html", data)
        else:
                return render(request, "empresa_1/promedios/promedio_home.html")





def homeFrecuenciaDiseño(request,id):
        if request.user.is_authenticated:
                empresa = Empresa.objects.filter(id_empresa = id)
                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.get(nombre = "Diseño y produccion") #trar solo la ID de la etapa "Extraccion materia prima"
                area = AreaEmpresa.objects.filter(id_empresa = id)
                            
                data = {
                'registros': registros,
                'etapa':etapa,
                'area' : area,
                'empresa':empresa,    

                }   
                return render(request,"diseñoProduccion/frecuencia/home_frecuencia.html", data)
        else:
                return render(request, "empresa_1/promedios/promedio_home.html")


def homeFrecuenciaLogistica(request,id):
        if request.user.is_authenticated:
                empresa = Empresa.objects.filter(id_empresa = id)
                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.get(nombre = "Logistica") #trar solo la ID de la etapa "Extraccion materia prima"
                area = AreaEmpresa.objects.filter(id_empresa = id)
                            
                data = {
                'registros': registros,
                'etapa':etapa,
                'area' : area,
                'empresa':empresa,    

                }   
                return render(request,"logistica/frecuencia/home_frecuencia.html", data)
        else:
                return render(request, "logistica/frecuencia/home_frecuencia.html")

def homeFrecuenciaCompra(request,id):
        if request.user.is_authenticated:
                empresa = Empresa.objects.filter(id_empresa = id)
                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.get(nombre = "Compra") #trar solo la ID de la etapa "Extraccion materia prima"
                area = AreaEmpresa.objects.filter(id_empresa = id)
                            
                data = {
                'registros': registros,
                'etapa':etapa,
                'area' : area,
                'empresa':empresa,    

                }   
                return render(request,"compra/frecuencia/home_frecuencia.html", data)
        else:
                return render(request, "compra/frecuencia/home_frecuencia.html")





def homeFrecuenciaUso(request,id):
        if request.user.is_authenticated:
                empresa = Empresa.objects.filter(id_empresa = id)
                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.get(nombre = "Uso consumo") #trar solo la ID de la etapa "Extraccion materia prima"
                area = AreaEmpresa.objects.filter(id_empresa = id)
                            
                data = {
                'registros': registros,
                'etapa':etapa,
                'area' : area,
                'empresa':empresa,    

                }   

                return render(request,"usoConsumo/frecuencia/home_frecuencia.html", data)
        else:
                return render(request, "compra/frecuencia/home_frecuencia.html")



def homeFrecuenciaFin(request,id):
        if request.user.is_authenticated:
                empresa = Empresa.objects.filter(id_empresa = id)
                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.get(nombre = "Fin de vida") #trar solo la ID de la etapa "Extraccion materia prima"
                area = AreaEmpresa.objects.filter(id_empresa = id)
                            
                data = {
                'registros': registros,
                'etapa':etapa,
                'area' : area,
                'empresa':empresa,    

                }   

                return render(request,"finVida/frecuencia/home_frecuencia.html", data)
        else:
                return render(request, "compra/frecuencia/home_frecuencia.html")




def promedioArea(request, id):
        area = AreaEmpresa.objects.filter(id_area = id)
        etapa = Etapa.objects.get(nombre = "Extraccion materia prima")
        registros = RegistroTrabajador.objects.filter(usuario=request.user)    
        b = 0
        empresa = 0
        for a in area:
                if b < 1 :
                        empresa_id = a.id_empresa_id
                        b = b + 1
        print(empresa_id)        
        empresa = Empresa.objects.filter(id_empresa = empresa_id )


        
        entradas = Entrada.objects.filter(id_area_id = id, etapa_id = etapa)
        salidas = Salida.objects.filter(id_area_id = id, etapa_id = etapa)
        oportunidades = Oportunidades.objects.filter(id_area_id = id, etapa_id = etapa)
        total_entradas = Entrada.objects.filter(id_area_id = id).count()
        total_salidas = Salida.objects.filter(id_area_id = id).count()
        total_oportunidades = Oportunidades.objects.filter(id_area_id = id).count()

        print("la id de la estapa es!!!!!!!!: ",etapa)
        

#/////////////// Entradas /////////////////////////////////////////////////////////////////////////////////////////////////////////
        lista_u = []
        plabra_es = []
        lista_t = [] 
        nombre_espa = ""

        for e in entradas:
                #print("las notas son: ",e.nombre) 
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                #print(result)
                #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_espa = e.nombre
                        #print(nombre_espa)
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_espa.split(separador, maximo_numero_de_separaciones)
                        plabra_es = plabra_es + separado_por_espacios
                                                
                else:
                        lista_u.append(e_nombre)
                        
        lista_t =  plabra_es + lista_u                       
        c = collections.Counter(lista_t) #crea un diccionario agrupando por palabras

        clave = c.keys()
        valor = c.values()
        cantidad_datos = c.items()

        # for clave, valor in cantidad_datos:
        #         print (clave , ": " , valor)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

#/////////////// salidas /////////////////////////////////////////////////////////////////////////////////////////////////////////
        lista_unita_salida = []
        plabra_espacio_salida = []
        lista_total_salida = [] 
        nombre_espa_con_espacio = ""

        for e in salidas:
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
                # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_espa_con_espacio = e.nombre
                
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_espa_con_espacio.split(separador, maximo_numero_de_separaciones)
                        plabra_espacio_salida = plabra_espacio_salida + separado_por_espacios
                                                
                else:
                        lista_unita_salida.append(e_nombre)

        lista_total_salida = plabra_espacio_salida + lista_unita_salida

        print(lista_total_salida)
        c_salidas = collections.Counter(lista_total_salida)
        print(c_salidas)


        clave_salidas = c_salidas.keys()
        valor_salidas = c_salidas.values()
        cantidad_datos_salidas = c_salidas.items()

        # for clave, valor in cantidad_datos:
        #         print (clave , ": " , valor)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

#/////////////// Oportunidades /////////////////////////////////////////////////////////////////////////////////////////////////////////

        lista_unita_opor = []                           #nota sin espacio en su nombre
        plabra_espacio_opor = []                        #nota con espacio en su nombre
        lista_total_opor = []                           #arrays con todas las notas
        nombre_con_espacio_opor = ""                       #variable donde se guarda la nota con espacios en su nombre deontro del for

        for e in oportunidades:
                
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
                # # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_con_espacio_opor = e.nombre
                
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_con_espacio_opor.split(separador, maximo_numero_de_separaciones)
                        plabra_espacio_opor = plabra_espacio_opor + separado_por_espacios
                                                
                else:
                        lista_unita_opor.append(e_nombre)

        lista_total_opor = plabra_espacio_opor + lista_unita_opor

        # print(lista_total_salida)
        c_oportunidad = collections.Counter(lista_total_opor)
        print("las oportunidades son:  ", c_oportunidad)


        clave_oportunidad = c_oportunidad.keys()
        valor_oportunidad = c_oportunidad.values()
        cantidad_datos_oportunidad = c_oportunidad.items()

        # for clave_oportunidad, valor_oportunidad in cantidad_datos:
        #         print (clave_oportunidad , ": " , valor_oportunidad)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 


        values = c.values()
        #print(c['harina'])
        p = c['harina']
        

        nota_masRepetida = ""
        b = 0
        for i in c:
                n = c[i]
                if b < n:
                        b = n 
                               
                if n == b :
                        nota_masRepetida = i
        #print("la nota mas repetida es: ", nota_masRepetida) 
       

        #///////////// Levenshetein ///////////////////////////
        import numpy as np
        # filas = len(lista_t)
        # columnas = len(lista_t)


        # A = np.zeros([filas,columnas])

        # for i in range(filas):
        #         for j in range(columnas):
        #                 a = i+1
        #                 x = j+1
        #                 A[i,j] = a+x
        # print(A)

        n_filas = len(lista_t)          #Cantidad de registros
        n_columnas = len(lista_t)       #Cantidad de registros
        
        A = np.zeros([n_filas, n_columnas]) #Creo un array vacio

        for i in range(n_filas):
                for j in range(n_columnas):
                        a = lista_t[i]
                        b = lista_t[j]
                        A[i,j] = round(jaro(a,b)*100)
                       
        #print(A)   
        # for line in A:
        #         c = map
        #         print (' '.join(map(str, line)))
        arr = A
        # for i in A:
        #         print (i)
       
        # for i in range(n_filas):
        #         canti_registros =  i
        #         columna = [fila[i] for fila in A] 

                #print(columna)    
        #print(canti_registros)
        #////////////////////////////////////////                     
        data = {
                'registros'                     :       registros,
                'area'                          :       area,
                'etapa'                         :       etapa,
                'empresa'                       :       empresa,
                'total_entradas'                :       total_entradas,
                'total_salidas'                 :       total_salidas,
                'total_oportunidades'           :       total_salidas,
                'nota_masRepetida'              :       nota_masRepetida,
                'lista_t'                       :       lista_t,

                #entradas
                'clave'                         :       clave,
                'valor'                         :       valor,
                'cantidad_datos'                :       cantidad_datos,

                #salidas
                'clave_salidas'                 :       clave_salidas,
                'valor_salidas'                 :       valor_salidas,
                'cantidad_datos_salidas'        :       cantidad_datos_salidas,

                #oportunidad
                'clave_oportunidad'             :       clave_oportunidad,
                'valor_oportunidad'             :       valor_oportunidad,
                'cantidad_datos_oportunidad'    :       cantidad_datos_oportunidad,

                'A':A,
                'arr'                           :       arr,
                'n_filas'                       :       n_filas,
                'n_columnas'                    :       n_columnas,
                #'canti_registros'               :       canti_registros,
                        
        }

        return render(request,'empresa_1/promedios/promedio.html', data)   



def frecuenciaDiseño(request, id):
        area = AreaEmpresa.objects.filter(id_area = id)
        etapa = Etapa.objects.get(nombre = "Diseño y produccion")
        registros = RegistroTrabajador.objects.filter(usuario=request.user)    
        b = 0
        empresa = 0
        for a in area:
                if b < 1 :
                        empresa_id = a.id_empresa_id
                        b = b + 1
        print(empresa_id)        

        empresa = Empresa.objects.filter(id_empresa = empresa_id )


        
        entradas = Entrada.objects.filter(id_area_id = id, etapa_id = etapa)
        salidas = Salida.objects.filter(id_area_id = id, etapa_id = etapa)
        oportunidades = Oportunidades.objects.filter(id_area_id = id, etapa_id = etapa)
        total_entradas = Entrada.objects.filter(id_area_id = id).count()
        total_salidas = Salida.objects.filter(id_area_id = id).count()
        total_oportunidades = Oportunidades.objects.filter(id_area_id = id).count()

        print("la id de la estapa es!!!!!!!!: ",etapa)
        

#/////////////// Entradas /////////////////////////////////////////////////////////////////////////////////////////////////////////
        lista_u = []
        plabra_es = []
        lista_t = [] 
        nombre_espa = ""

        for e in entradas:
                #print("las notas son: ",e.nombre) 
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                #print(result)
                #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_espa = e.nombre
                        #print(nombre_espa)
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_espa.split(separador, maximo_numero_de_separaciones)
                        plabra_es = plabra_es + separado_por_espacios
                                                
                else:
                        lista_u.append(e_nombre)
                        
        lista_t =  plabra_es + lista_u                       
        c = collections.Counter(lista_t) #crea un diccionario agrupando por palabras

        clave = c.keys()
        valor = c.values()
        cantidad_datos = c.items()

        # for clave, valor in cantidad_datos:
        #         print (clave , ": " , valor)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

#/////////////// salidas /////////////////////////////////////////////////////////////////////////////////////////////////////////
        lista_unita_salida = []
        plabra_espacio_salida = []
        lista_total_salida = [] 
        nombre_espa_con_espacio = ""

        for e in salidas:
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
                # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_espa_con_espacio = e.nombre
                
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_espa_con_espacio.split(separador, maximo_numero_de_separaciones)
                        plabra_espacio_salida = plabra_espacio_salida + separado_por_espacios
                                                
                else:
                        lista_unita_salida.append(e_nombre)

        lista_total_salida = plabra_espacio_salida + lista_unita_salida

        print(lista_total_salida)
        c_salidas = collections.Counter(lista_total_salida)
        print(c_salidas)


        clave_salidas = c_salidas.keys()
        valor_salidas = c_salidas.values()
        cantidad_datos_salidas = c_salidas.items()

        # for clave, valor in cantidad_datos:
        #         print (clave , ": " , valor)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

#/////////////// Oportunidades /////////////////////////////////////////////////////////////////////////////////////////////////////////

        lista_unita_opor = []                           #nota sin espacio en su nombre
        plabra_espacio_opor = []                        #nota con espacio en su nombre
        lista_total_opor = []                           #arrays con todas las notas
        nombre_con_espacio_opor = ""                       #variable donde se guarda la nota con espacios en su nombre deontro del for

        for e in oportunidades:
                
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
                # # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_con_espacio_opor = e.nombre
                
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_con_espacio_opor.split(separador, maximo_numero_de_separaciones)
                        plabra_espacio_opor = plabra_espacio_opor + separado_por_espacios
                                                
                else:
                        lista_unita_opor.append(e_nombre)

        lista_total_opor = plabra_espacio_opor + lista_unita_opor

        # print(lista_total_salida)
        c_oportunidad = collections.Counter(lista_total_opor)
        print("las oportunidades son:  ", c_oportunidad)


        clave_oportunidad = c_oportunidad.keys()
        valor_oportunidad = c_oportunidad.values()
        cantidad_datos_oportunidad = c_oportunidad.items()

        # for clave_oportunidad, valor_oportunidad in cantidad_datos:
        #         print (clave_oportunidad , ": " , valor_oportunidad)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

        nota_masRepetida = ""
        b = 0
        for i in c:
                n = c[i]
                if b < n:
                        b = n 
                               
                if n == b :
                        nota_masRepetida = i   

        #///////////// Levenshetein ///////////////////////////
        import numpy as np
        

        n_filas = len(lista_t)          #Cantidad de registros
        n_columnas = len(lista_t)       #Cantidad de registros
        
        A = np.zeros([n_filas, n_columnas]) #Creo un array vacio

        for i in range(n_filas):
                for j in range(n_columnas):
                        a = lista_t[i]
                        b = lista_t[j]
                        A[i,j] = round(jaro(a,b)*100)
                       
        arr = A
        
        #////////////////////////////////////////                     
        data = {
                'registros'                     :       registros,
                'area'                          :       area,
                'etapa'                         :       etapa,
                'empresa'                       :       empresa,
                'total_entradas'                :       total_entradas,
                'total_salidas'                 :       total_salidas,
                'total_oportunidades'           :       total_salidas,
                'nota_masRepetida'              :       nota_masRepetida,
                'lista_t'                       :       lista_t,

                #entradas
                'clave'                         :       clave,
                'valor'                         :       valor,
                'cantidad_datos'                :       cantidad_datos,

                #salidas
                'clave_salidas'                 :       clave_salidas,
                'valor_salidas'                 :       valor_salidas,
                'cantidad_datos_salidas'        :       cantidad_datos_salidas,

                #oportunidad
                'clave_oportunidad'             :       clave_oportunidad,
                'valor_oportunidad'             :       valor_oportunidad,
                'cantidad_datos_oportunidad'    :       cantidad_datos_oportunidad,

                'A':A,
                'arr'                           :       arr,
                'n_filas'                       :       n_filas,
                'n_columnas'                    :       n_columnas,
                #'canti_registros'               :       canti_registros,
                        
        }

        return render(request,'diseñoProduccion/frecuencia/tablas_frecuencia.html', data)   






def frecuenciaLogistica(request, id):
        area = AreaEmpresa.objects.filter(id_area = id)
        etapa = Etapa.objects.get(nombre = "Logistica")
        registros = RegistroTrabajador.objects.filter(usuario=request.user)    
        b = 0
        empresa = 0
        for a in area:
                if b < 1 :
                        empresa_id = a.id_empresa_id
                        b = b + 1
        print(empresa_id)        

        empresa = Empresa.objects.filter(id_empresa = empresa_id )


        
        entradas = Entrada.objects.filter(id_area_id = id, etapa_id = etapa)
        salidas = Salida.objects.filter(id_area_id = id, etapa_id = etapa)
        oportunidades = Oportunidades.objects.filter(id_area_id = id, etapa_id = etapa)
        total_entradas = Entrada.objects.filter(id_area_id = id).count()
        total_salidas = Salida.objects.filter(id_area_id = id).count()
        total_oportunidades = Oportunidades.objects.filter(id_area_id = id).count()

        print("la id de la estapa es!!!!!!!!: ",etapa)
        

#/////////////// Entradas /////////////////////////////////////////////////////////////////////////////////////////////////////////
        lista_u = []
        plabra_es = []
        lista_t = [] 
        nombre_espa = ""

        for e in entradas:
                #print("las notas son: ",e.nombre) 
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                #print(result)
                #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_espa = e.nombre
                        #print(nombre_espa)
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_espa.split(separador, maximo_numero_de_separaciones)
                        plabra_es = plabra_es + separado_por_espacios
                                                
                else:
                        lista_u.append(e_nombre)
                        
        lista_t =  plabra_es + lista_u                       
        c = collections.Counter(lista_t) #crea un diccionario agrupando por palabras

        clave = c.keys()
        valor = c.values()
        cantidad_datos = c.items()

        # for clave, valor in cantidad_datos:
        #         print (clave , ": " , valor)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

#/////////////// salidas /////////////////////////////////////////////////////////////////////////////////////////////////////////
        lista_unita_salida = []
        plabra_espacio_salida = []
        lista_total_salida = [] 
        nombre_espa_con_espacio = ""

        for e in salidas:
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
                # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_espa_con_espacio = e.nombre
                
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_espa_con_espacio.split(separador, maximo_numero_de_separaciones)
                        plabra_espacio_salida = plabra_espacio_salida + separado_por_espacios
                                                
                else:
                        lista_unita_salida.append(e_nombre)

        lista_total_salida = plabra_espacio_salida + lista_unita_salida

        print(lista_total_salida)
        c_salidas = collections.Counter(lista_total_salida)
        print(c_salidas)


        clave_salidas = c_salidas.keys()
        valor_salidas = c_salidas.values()
        cantidad_datos_salidas = c_salidas.items()

        # for clave, valor in cantidad_datos:
        #         print (clave , ": " , valor)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

#/////////////// Oportunidades /////////////////////////////////////////////////////////////////////////////////////////////////////////

        lista_unita_opor = []                           #nota sin espacio en su nombre
        plabra_espacio_opor = []                        #nota con espacio en su nombre
        lista_total_opor = []                           #arrays con todas las notas
        nombre_con_espacio_opor = ""                       #variable donde se guarda la nota con espacios en su nombre deontro del for

        for e in oportunidades:
                
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
                # # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_con_espacio_opor = e.nombre
                
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_con_espacio_opor.split(separador, maximo_numero_de_separaciones)
                        plabra_espacio_opor = plabra_espacio_opor + separado_por_espacios
                                                
                else:
                        lista_unita_opor.append(e_nombre)

        lista_total_opor = plabra_espacio_opor + lista_unita_opor

        # print(lista_total_salida)
        c_oportunidad = collections.Counter(lista_total_opor)
        print("las oportunidades son:  ", c_oportunidad)


        clave_oportunidad = c_oportunidad.keys()
        valor_oportunidad = c_oportunidad.values()
        cantidad_datos_oportunidad = c_oportunidad.items()

        # for clave_oportunidad, valor_oportunidad in cantidad_datos:
        #         print (clave_oportunidad , ": " , valor_oportunidad)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

        nota_masRepetida = ""
        b = 0
        for i in c:
                n = c[i]
                if b < n:
                        b = n 
                               
                if n == b :
                        nota_masRepetida = i   

        #///////////// Levenshetein ///////////////////////////
        import numpy as np
        

        n_filas = len(lista_t)          #Cantidad de registros
        n_columnas = len(lista_t)       #Cantidad de registros
        
        A = np.zeros([n_filas, n_columnas]) #Creo un array vacio

        for i in range(n_filas):
                for j in range(n_columnas):
                        a = lista_t[i]
                        b = lista_t[j]
                        A[i,j] = round(jaro(a,b)*100)
                       
        arr = A
        
        #////////////////////////////////////////                     
        data = {
                'registros'                     :       registros,
                'area'                          :       area,
                'etapa'                         :       etapa,
                'empresa'                       :       empresa,
                'total_entradas'                :       total_entradas,
                'total_salidas'                 :       total_salidas,
                'total_oportunidades'           :       total_salidas,
                'nota_masRepetida'              :       nota_masRepetida,
                'lista_t'                       :       lista_t,

                #entradas
                'clave'                         :       clave,
                'valor'                         :       valor,
                'cantidad_datos'                :       cantidad_datos,

                #salidas
                'clave_salidas'                 :       clave_salidas,
                'valor_salidas'                 :       valor_salidas,
                'cantidad_datos_salidas'        :       cantidad_datos_salidas,

                #oportunidad
                'clave_oportunidad'             :       clave_oportunidad,
                'valor_oportunidad'             :       valor_oportunidad,
                'cantidad_datos_oportunidad'    :       cantidad_datos_oportunidad,

                'A':A,
                'arr'                           :       arr,
                'n_filas'                       :       n_filas,
                'n_columnas'                    :       n_columnas,
                #'canti_registros'               :       canti_registros,
                        
        }

        return render(request,'logistica/frecuencia/tablas_frecuencia.html', data)   



def frecuenciaCompra(request, id):
        area = AreaEmpresa.objects.filter(id_area = id)
        etapa = Etapa.objects.get(nombre = "Compra")
        registros = RegistroTrabajador.objects.filter(usuario=request.user)    
        b = 0
        empresa = 0
        for a in area:
                if b < 1 :
                        empresa_id = a.id_empresa_id
                        b = b + 1
        print(empresa_id)        

        empresa = Empresa.objects.filter(id_empresa = empresa_id )


        
        entradas = Entrada.objects.filter(id_area_id = id, etapa_id = etapa)
        salidas = Salida.objects.filter(id_area_id = id, etapa_id = etapa)
        oportunidades = Oportunidades.objects.filter(id_area_id = id, etapa_id = etapa)
        total_entradas = Entrada.objects.filter(id_area_id = id).count()
        total_salidas = Salida.objects.filter(id_area_id = id).count()
        total_oportunidades = Oportunidades.objects.filter(id_area_id = id).count()

        print("la id de la estapa es!!!!!!!!: ",etapa)
        

#/////////////// Entradas /////////////////////////////////////////////////////////////////////////////////////////////////////////
        lista_u = []
        plabra_es = []
        lista_t = [] 
        nombre_espa = ""

        for e in entradas:
                #print("las notas son: ",e.nombre) 
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                #print(result)
                #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_espa = e.nombre
                        #print(nombre_espa)
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_espa.split(separador, maximo_numero_de_separaciones)
                        plabra_es = plabra_es + separado_por_espacios
                                                
                else:
                        lista_u.append(e_nombre)
                        
        lista_t =  plabra_es + lista_u                       
        c = collections.Counter(lista_t) #crea un diccionario agrupando por palabras

        clave = c.keys()
        valor = c.values()
        cantidad_datos = c.items()

        # for clave, valor in cantidad_datos:
        #         print (clave , ": " , valor)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

#/////////////// salidas /////////////////////////////////////////////////////////////////////////////////////////////////////////
        lista_unita_salida = []
        plabra_espacio_salida = []
        lista_total_salida = [] 
        nombre_espa_con_espacio = ""

        for e in salidas:
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
                # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_espa_con_espacio = e.nombre
                
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_espa_con_espacio.split(separador, maximo_numero_de_separaciones)
                        plabra_espacio_salida = plabra_espacio_salida + separado_por_espacios
                                                
                else:
                        lista_unita_salida.append(e_nombre)

        lista_total_salida = plabra_espacio_salida + lista_unita_salida

        print(lista_total_salida)
        c_salidas = collections.Counter(lista_total_salida)
        print(c_salidas)


        clave_salidas = c_salidas.keys()
        valor_salidas = c_salidas.values()
        cantidad_datos_salidas = c_salidas.items()

        # for clave, valor in cantidad_datos:
        #         print (clave , ": " , valor)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

#/////////////// Oportunidades /////////////////////////////////////////////////////////////////////////////////////////////////////////

        lista_unita_opor = []                           #nota sin espacio en su nombre
        plabra_espacio_opor = []                        #nota con espacio en su nombre
        lista_total_opor = []                           #arrays con todas las notas
        nombre_con_espacio_opor = ""                       #variable donde se guarda la nota con espacios en su nombre deontro del for

        for e in oportunidades:
                
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
                # # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_con_espacio_opor = e.nombre
                
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_con_espacio_opor.split(separador, maximo_numero_de_separaciones)
                        plabra_espacio_opor = plabra_espacio_opor + separado_por_espacios
                                                
                else:
                        lista_unita_opor.append(e_nombre)

        lista_total_opor = plabra_espacio_opor + lista_unita_opor

        # print(lista_total_salida)
        c_oportunidad = collections.Counter(lista_total_opor)
        print("las oportunidades son:  ", c_oportunidad)


        clave_oportunidad = c_oportunidad.keys()
        valor_oportunidad = c_oportunidad.values()
        cantidad_datos_oportunidad = c_oportunidad.items()

        # for clave_oportunidad, valor_oportunidad in cantidad_datos:
        #         print (clave_oportunidad , ": " , valor_oportunidad)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

        nota_masRepetida = ""
        b = 0
        for i in c:
                n = c[i]
                if b < n:
                        b = n 
                               
                if n == b :
                        nota_masRepetida = i   

        #///////////// Levenshetein ///////////////////////////
        import numpy as np
        

        n_filas = len(lista_t)          #Cantidad de registros
        n_columnas = len(lista_t)       #Cantidad de registros
        
        A = np.zeros([n_filas, n_columnas]) #Creo un array vacio

        for i in range(n_filas):
                for j in range(n_columnas):
                        a = lista_t[i]
                        b = lista_t[j]
                        A[i,j] = round(jaro(a,b)*100)
                       
        arr = A
        
        #////////////////////////////////////////                     
        data = {
                'registros'                     :       registros,
                'area'                          :       area,
                'etapa'                         :       etapa,
                'empresa'                       :       empresa,
                'total_entradas'                :       total_entradas,
                'total_salidas'                 :       total_salidas,
                'total_oportunidades'           :       total_salidas,
                'nota_masRepetida'              :       nota_masRepetida,
                'lista_t'                       :       lista_t,

                #entradas
                'clave'                         :       clave,
                'valor'                         :       valor,
                'cantidad_datos'                :       cantidad_datos,

                #salidas
                'clave_salidas'                 :       clave_salidas,
                'valor_salidas'                 :       valor_salidas,
                'cantidad_datos_salidas'        :       cantidad_datos_salidas,

                #oportunidad
                'clave_oportunidad'             :       clave_oportunidad,
                'valor_oportunidad'             :       valor_oportunidad,
                'cantidad_datos_oportunidad'    :       cantidad_datos_oportunidad,

                'A':A,
                'arr'                           :       arr,
                'n_filas'                       :       n_filas,
                'n_columnas'                    :       n_columnas,
                #'canti_registros'               :       canti_registros,
                        
        }

        return render(request,'compra/frecuencia/tablas_frecuencia.html', data)  



def frecuenciaUso(request, id):
        area = AreaEmpresa.objects.filter(id_area = id)
        etapa = Etapa.objects.get(nombre = "Uso consumo")
        registros = RegistroTrabajador.objects.filter(usuario=request.user)    
        b = 0
        empresa = 0
        for a in area:
                if b < 1 :
                        empresa_id = a.id_empresa_id
                        b = b + 1
        print(empresa_id)        

        empresa = Empresa.objects.filter(id_empresa = empresa_id )


        
        entradas = Entrada.objects.filter(id_area_id = id, etapa_id = etapa)
        salidas = Salida.objects.filter(id_area_id = id, etapa_id = etapa)
        oportunidades = Oportunidades.objects.filter(id_area_id = id, etapa_id = etapa)
        total_entradas = Entrada.objects.filter(id_area_id = id).count()
        total_salidas = Salida.objects.filter(id_area_id = id).count()
        total_oportunidades = Oportunidades.objects.filter(id_area_id = id).count()

        print("la id de la estapa es!!!!!!!!: ",etapa)
        

#/////////////// Entradas /////////////////////////////////////////////////////////////////////////////////////////////////////////
        lista_u = []
        plabra_es = []
        lista_t = [] 
        nombre_espa = ""

        for e in entradas:
                #print("las notas son: ",e.nombre) 
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                #print(result)
                #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_espa = e.nombre
                        #print(nombre_espa)
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_espa.split(separador, maximo_numero_de_separaciones)
                        plabra_es = plabra_es + separado_por_espacios
                                                
                else:
                        lista_u.append(e_nombre)
                        
        lista_t =  plabra_es + lista_u                       
        c = collections.Counter(lista_t) #crea un diccionario agrupando por palabras

        clave = c.keys()
        valor = c.values()
        cantidad_datos = c.items()

        # for clave, valor in cantidad_datos:
        #         print (clave , ": " , valor)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

#/////////////// salidas /////////////////////////////////////////////////////////////////////////////////////////////////////////
        lista_unita_salida = []
        plabra_espacio_salida = []
        lista_total_salida = [] 
        nombre_espa_con_espacio = ""

        for e in salidas:
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
                # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_espa_con_espacio = e.nombre
                
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_espa_con_espacio.split(separador, maximo_numero_de_separaciones)
                        plabra_espacio_salida = plabra_espacio_salida + separado_por_espacios
                                                
                else:
                        lista_unita_salida.append(e_nombre)

        lista_total_salida = plabra_espacio_salida + lista_unita_salida

        print(lista_total_salida)
        c_salidas = collections.Counter(lista_total_salida)
        print(c_salidas)


        clave_salidas = c_salidas.keys()
        valor_salidas = c_salidas.values()
        cantidad_datos_salidas = c_salidas.items()

        # for clave, valor in cantidad_datos:
        #         print (clave , ": " , valor)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

#/////////////// Oportunidades /////////////////////////////////////////////////////////////////////////////////////////////////////////

        lista_unita_opor = []                           #nota sin espacio en su nombre
        plabra_espacio_opor = []                        #nota con espacio en su nombre
        lista_total_opor = []                           #arrays con todas las notas
        nombre_con_espacio_opor = ""                       #variable donde se guarda la nota con espacios en su nombre deontro del for

        for e in oportunidades:
                
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
                # # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_con_espacio_opor = e.nombre
                
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_con_espacio_opor.split(separador, maximo_numero_de_separaciones)
                        plabra_espacio_opor = plabra_espacio_opor + separado_por_espacios
                                                
                else:
                        lista_unita_opor.append(e_nombre)

        lista_total_opor = plabra_espacio_opor + lista_unita_opor

        # print(lista_total_salida)
        c_oportunidad = collections.Counter(lista_total_opor)
        print("las oportunidades son:  ", c_oportunidad)


        clave_oportunidad = c_oportunidad.keys()
        valor_oportunidad = c_oportunidad.values()
        cantidad_datos_oportunidad = c_oportunidad.items()

        # for clave_oportunidad, valor_oportunidad in cantidad_datos:
        #         print (clave_oportunidad , ": " , valor_oportunidad)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

        nota_masRepetida = ""
        b = 0
        for i in c:
                n = c[i]
                if b < n:
                        b = n 
                               
                if n == b :
                        nota_masRepetida = i   

        #///////////// Levenshetein ///////////////////////////
        import numpy as np
        

        n_filas = len(lista_t)          #Cantidad de registros
        n_columnas = len(lista_t)       #Cantidad de registros
        
        A = np.zeros([n_filas, n_columnas]) #Creo un array vacio

        for i in range(n_filas):
                for j in range(n_columnas):
                        a = lista_t[i]
                        b = lista_t[j]
                        A[i,j] = round(jaro(a,b)*100)
                       
        arr = A
        
        #////////////////////////////////////////                     
        data = {
                'registros'                     :       registros,
                'area'                          :       area,
                'etapa'                         :       etapa,
                'empresa'                       :       empresa,
                'total_entradas'                :       total_entradas,
                'total_salidas'                 :       total_salidas,
                'total_oportunidades'           :       total_salidas,
                'nota_masRepetida'              :       nota_masRepetida,
                'lista_t'                       :       lista_t,

                #entradas
                'clave'                         :       clave,
                'valor'                         :       valor,
                'cantidad_datos'                :       cantidad_datos,

                #salidas
                'clave_salidas'                 :       clave_salidas,
                'valor_salidas'                 :       valor_salidas,
                'cantidad_datos_salidas'        :       cantidad_datos_salidas,

                #oportunidad
                'clave_oportunidad'             :       clave_oportunidad,
                'valor_oportunidad'             :       valor_oportunidad,
                'cantidad_datos_oportunidad'    :       cantidad_datos_oportunidad,

                'A':A,
                'arr'                           :       arr,
                'n_filas'                       :       n_filas,
                'n_columnas'                    :       n_columnas,
                #'canti_registros'               :       canti_registros,
                        
        }

        return render(request,'usoConsumo/frecuencia/tablas_frecuencia.html', data)





def frecuenciaFin(request, id):
        area = AreaEmpresa.objects.filter(id_area = id)
        etapa = Etapa.objects.get(nombre = "Uso consumo")
        registros = RegistroTrabajador.objects.filter(usuario=request.user)    
        b = 0
        empresa = 0
        for a in area:
                if b < 1 :
                        empresa_id = a.id_empresa_id
                        b = b + 1
        print(empresa_id)        

        empresa = Empresa.objects.filter(id_empresa = empresa_id )


        
        entradas = Entrada.objects.filter(id_area_id = id, etapa_id = etapa)
        salidas = Salida.objects.filter(id_area_id = id, etapa_id = etapa)
        oportunidades = Oportunidades.objects.filter(id_area_id = id, etapa_id = etapa)
        total_entradas = Entrada.objects.filter(id_area_id = id).count()
        total_salidas = Salida.objects.filter(id_area_id = id).count()
        total_oportunidades = Oportunidades.objects.filter(id_area_id = id).count()

        print("la id de la estapa es!!!!!!!!: ",etapa)
        

#/////////////// Entradas /////////////////////////////////////////////////////////////////////////////////////////////////////////
        lista_u = []
        plabra_es = []
        lista_t = [] 
        nombre_espa = ""

        for e in entradas:
                #print("las notas son: ",e.nombre) 
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                #print(result)
                #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_espa = e.nombre
                        #print(nombre_espa)
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_espa.split(separador, maximo_numero_de_separaciones)
                        plabra_es = plabra_es + separado_por_espacios
                                                
                else:
                        lista_u.append(e_nombre)
                        
        lista_t =  plabra_es + lista_u                       
        c = collections.Counter(lista_t) #crea un diccionario agrupando por palabras

        clave = c.keys()
        valor = c.values()
        cantidad_datos = c.items()

        # for clave, valor in cantidad_datos:
        #         print (clave , ": " , valor)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

#/////////////// salidas /////////////////////////////////////////////////////////////////////////////////////////////////////////
        lista_unita_salida = []
        plabra_espacio_salida = []
        lista_total_salida = [] 
        nombre_espa_con_espacio = ""

        for e in salidas:
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
                # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_espa_con_espacio = e.nombre
                
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_espa_con_espacio.split(separador, maximo_numero_de_separaciones)
                        plabra_espacio_salida = plabra_espacio_salida + separado_por_espacios
                                                
                else:
                        lista_unita_salida.append(e_nombre)

        lista_total_salida = plabra_espacio_salida + lista_unita_salida

        print(lista_total_salida)
        c_salidas = collections.Counter(lista_total_salida)
        print(c_salidas)


        clave_salidas = c_salidas.keys()
        valor_salidas = c_salidas.values()
        cantidad_datos_salidas = c_salidas.items()

        # for clave, valor in cantidad_datos:
        #         print (clave , ": " , valor)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

#/////////////// Oportunidades /////////////////////////////////////////////////////////////////////////////////////////////////////////

        lista_unita_opor = []                           #nota sin espacio en su nombre
        plabra_espacio_opor = []                        #nota con espacio en su nombre
        lista_total_opor = []                           #arrays con todas las notas
        nombre_con_espacio_opor = ""                       #variable donde se guarda la nota con espacios en su nombre deontro del for

        for e in oportunidades:
                
                e_nombre = e.nombre
                result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
                # # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
                if result > 1 :
                        nombre_con_espacio_opor = e.nombre
                
                        separador = " "
                        maximo_numero_de_separaciones = 2
                        separado_por_espacios = nombre_con_espacio_opor.split(separador, maximo_numero_de_separaciones)
                        plabra_espacio_opor = plabra_espacio_opor + separado_por_espacios
                                                
                else:
                        lista_unita_opor.append(e_nombre)

        lista_total_opor = plabra_espacio_opor + lista_unita_opor

        # print(lista_total_salida)
        c_oportunidad = collections.Counter(lista_total_opor)
        print("las oportunidades son:  ", c_oportunidad)


        clave_oportunidad = c_oportunidad.keys()
        valor_oportunidad = c_oportunidad.values()
        cantidad_datos_oportunidad = c_oportunidad.items()

        # for clave_oportunidad, valor_oportunidad in cantidad_datos:
        #         print (clave_oportunidad , ": " , valor_oportunidad)
 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

        nota_masRepetida = ""
        b = 0
        for i in c:
                n = c[i]
                if b < n:
                        b = n 
                               
                if n == b :
                        nota_masRepetida = i   

        #///////////// Levenshetein ///////////////////////////
        import numpy as np
        

        n_filas = len(lista_t)          #Cantidad de registros
        n_columnas = len(lista_t)       #Cantidad de registros
        
        A = np.zeros([n_filas, n_columnas]) #Creo un array vacio

        for i in range(n_filas):
                for j in range(n_columnas):
                        a = lista_t[i]
                        b = lista_t[j]
                        A[i,j] = round(jaro(a,b)*100)
                       
        arr = A
        
        #////////////////////////////////////////                     
        data = {
                'registros'                     :       registros,
                'area'                          :       area,
                'etapa'                         :       etapa,
                'empresa'                       :       empresa,
                'total_entradas'                :       total_entradas,
                'total_salidas'                 :       total_salidas,
                'total_oportunidades'           :       total_salidas,
                'nota_masRepetida'              :       nota_masRepetida,
                'lista_t'                       :       lista_t,

                #entradas
                'clave'                         :       clave,
                'valor'                         :       valor,
                'cantidad_datos'                :       cantidad_datos,

                #salidas
                'clave_salidas'                 :       clave_salidas,
                'valor_salidas'                 :       valor_salidas,
                'cantidad_datos_salidas'        :       cantidad_datos_salidas,

                #oportunidad
                'clave_oportunidad'             :       clave_oportunidad,
                'valor_oportunidad'             :       valor_oportunidad,
                'cantidad_datos_oportunidad'    :       cantidad_datos_oportunidad,

                'A':A,
                'arr'                           :       arr,
                'n_filas'                       :       n_filas,
                'n_columnas'                    :       n_columnas,
                #'canti_registros'               :       canti_registros,
                        
        }

        return render(request,'finVida/frecuencia/tablas_frecuencia.html', data)


# def promedioArea(request, id):
#         etapa = Etapa.objects.get(nombre = "Extraccion materia prima") 
#         registros = RegistroTrabajador.objects.filter(usuario=request.user)
#         empresas = Empresa.objects.all()
#         empresa = Empresa.objects.filter(id_empresa = id)
#         area = AreaEmpresa.objects.filter(id_empresa = id)
#         empresaArea = RegistroTrabajador.objects.all()

#         entradas = Entrada.objects.filter(id_area_id = id, etapa_id = etapa)
#         salidas = Salida.objects.filter(id_area_id = id, etapa_id = etapa)
#         oportunidades = Oportunidades.objects.filter(id_area_id = id, etapa_id = etapa)
#         total_entradas = Entrada.objects.filter(id_area_id = id).count()
#         total_salidas = Salida.objects.filter(id_area_id = id).count()
#         total_oportunidades = Oportunidades.objects.filter(id_area_id = id).count()

#         print("la id de la estapa es!!!!!!!!: ",etapa)
        

# #/////////////// Entradas /////////////////////////////////////////////////////////////////////////////////////////////////////////
#         lista_u = []
#         plabra_es = []
#         lista_t = [] 
#         nombre_espa = ""

#         for e in entradas:
#                 #print("las notas son: ",e.nombre) 
#                 e_nombre = e.nombre
#                 result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
#                 #print(result)
#                 #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
#                 if result > 1 :
#                         nombre_espa = e.nombre
#                         #print(nombre_espa)
#                         separador = " "
#                         maximo_numero_de_separaciones = 2
#                         separado_por_espacios = nombre_espa.split(separador, maximo_numero_de_separaciones)
#                         plabra_es = plabra_es + separado_por_espacios
                                                
#                 else:
#                         lista_u.append(e_nombre)
                        
#         lista_t =  plabra_es + lista_u                       
#         c = collections.Counter(lista_t) #crea un diccionario agrupando por palabras

#         clave = c.keys()
#         valor = c.values()
#         cantidad_datos = c.items()

#         # for clave, valor in cantidad_datos:
#         #         print (clave , ": " , valor)
 
# #//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

# #/////////////// salidas /////////////////////////////////////////////////////////////////////////////////////////////////////////
#         lista_unita_salida = []
#         plabra_espacio_salida = []
#         lista_total_salida = [] 
#         nombre_espa_con_espacio = ""

#         for e in salidas:
#                 e_nombre = e.nombre
#                 result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
#                 # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
#                 if result > 1 :
#                         nombre_espa_con_espacio = e.nombre
                
#                         separador = " "
#                         maximo_numero_de_separaciones = 2
#                         separado_por_espacios = nombre_espa_con_espacio.split(separador, maximo_numero_de_separaciones)
#                         plabra_espacio_salida = plabra_espacio_salida + separado_por_espacios
                                                
#                 else:
#                         lista_unita_salida.append(e_nombre)

#         lista_total_salida = plabra_espacio_salida + lista_unita_salida

#         print(lista_total_salida)
#         c_salidas = collections.Counter(lista_total_salida)
#         print(c_salidas)


#         clave_salidas = c_salidas.keys()
#         valor_salidas = c_salidas.values()
#         cantidad_datos_salidas = c_salidas.items()

#         # for clave, valor in cantidad_datos:
#         #         print (clave , ": " , valor)
 
# #//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

# #/////////////// Oportunidades /////////////////////////////////////////////////////////////////////////////////////////////////////////

#         lista_unita_opor = []                           #nota sin espacio en su nombre
#         plabra_espacio_opor = []                        #nota con espacio en su nombre
#         lista_total_opor = []                           #arrays con todas las notas
#         nombre_con_espacio_opor = ""                       #variable donde se guarda la nota con espacios en su nombre deontro del for

#         for e in oportunidades:
                
#                 e_nombre = e.nombre
#                 result = len(e_nombre.split()) #cuanta las palabras que tiene cada registro
                
#                 # # #si el registro tiene mas de un 1 palabra guardalo en la variable nombre_espa
#                 if result > 1 :
#                         nombre_con_espacio_opor = e.nombre
                
#                         separador = " "
#                         maximo_numero_de_separaciones = 2
#                         separado_por_espacios = nombre_con_espacio_opor.split(separador, maximo_numero_de_separaciones)
#                         plabra_espacio_opor = plabra_espacio_opor + separado_por_espacios
                                                
#                 else:
#                         lista_unita_opor.append(e_nombre)

#         lista_total_opor = plabra_espacio_opor + lista_unita_opor

#         # print(lista_total_salida)
#         c_oportunidad = collections.Counter(lista_total_opor)
#         print("las oportunidades son:  ", c_oportunidad)


#         clave_oportunidad = c_oportunidad.keys()
#         valor_oportunidad = c_oportunidad.values()
#         cantidad_datos_oportunidad = c_oportunidad.items()

#         # for clave_oportunidad, valor_oportunidad in cantidad_datos:
#         #         print (clave_oportunidad , ": " , valor_oportunidad)
 
# #//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 


#         values = c.values()
#         #print(c['harina'])
#         p = c['harina']
        

#         nota_masRepetida = ""
#         b = 0
#         for i in c:
#                 n = c[i]
#                 if b < n:
#                         b = n 
                               
#                 if n == b :
#                         nota_masRepetida = i
#         #print("la nota mas repetida es: ", nota_masRepetida) 
       

#         #///////////// Levenshetein ///////////////////////////
#         import numpy as np
#         # filas = len(lista_t)
#         # columnas = len(lista_t)


#         # A = np.zeros([filas,columnas])

#         # for i in range(filas):
#         #         for j in range(columnas):
#         #                 a = i+1
#         #                 x = j+1
#         #                 A[i,j] = a+x
#         # print(A)

#         n_filas = len(lista_t)          #Cantidad de registros
#         n_columnas = len(lista_t)       #Cantidad de registros
        
#         A = np.zeros([n_filas, n_columnas]) #Creo un array vacio

#         for i in range(n_filas):
#                 for j in range(n_columnas):
#                         a = lista_t[i]
#                         b = lista_t[j]
#                         A[i,j] = round(jaro(a,b)*100)
                       
#         #print(A)   
#         # for line in A:
#         #         c = map
#         #         print (' '.join(map(str, line)))
#         arr = A
#         # for i in A:
#         #         print (i)
       
#         # for i in range(n_filas):
#         #         canti_registros =  i
#         #         columna = [fila[i] for fila in A] 

#                 #print(columna)    
#         #print(canti_registros)
#         #////////////////////////////////////////                

#         data = {

                
#                 'registros'                     :       registros,
#                 'empresas'                      :       empresas,
#                 'empresa'                       :       empresa,
#                 'total_entradas'                :       total_entradas,
#                 'total_salidas'                 :       total_salidas,
#                 'total_oportunidades'           :       total_salidas,
#                 'nota_masRepetida'              :       nota_masRepetida,
#                 'lista_t'                       :       lista_t,

#                 #entradas
#                 'clave'                         :       clave,
#                 'valor'                         :       valor,
#                 'cantidad_datos'                :       cantidad_datos,

#                 #salidas
#                 'clave_salidas'                 :       clave_salidas,
#                 'valor_salidas'                 :       valor_salidas,
#                 'cantidad_datos_salidas'        :       cantidad_datos_salidas,

#                 #oportunidad
#                 'clave_oportunidad'             :       clave_oportunidad,
#                 'valor_oportunidad'             :       valor_oportunidad,
#                 'cantidad_datos_oportunidad'    :       cantidad_datos_oportunidad,

#                 'A':A,
#                 'arr'                           :       arr,
#                 'n_filas'                       :       n_filas,
#                 'n_columnas'                    :       n_columnas,
#                 #'canti_registros'               :       canti_registros,
#                 'area'                          :       area,
#                 "empresaArea"                   :       empresaArea
 
#         }

#         return render(request, 'empresa_1/promedios/promedio.html', data)


def entradasExtraccion(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Extraccion materia prima")
                entradas = Entrada.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'extraccion/entrada/tabla_entrada.html', data)
        else:
                return render(request, 'extraccion/entrada/tabla_entrada.html')

def SalidasExtraccion(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Extraccion materia prima")
                entradas = Salida.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'extraccion/salida/tabla_salida.html', data)
        else:
                return render(request, 'extraccion/salida/tabla_salida.html')
        

def OportunidadExtraccion(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Extraccion materia prima")
                entradas = Oportunidades.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'extraccion/oportunidad/tabla_oportunidad.html', data)
        else:
                return render(request, 'extraccion/oportunidad/tabla_oportunidad.html')



# Diseño y gestion     

def EntradaDiseño(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Diseño y produccion")
                entradas = Entrada.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'diseñoProduccion/entrada/tabla_entrada.html', data)
        else:
                return render(request, 'diseñoProduccion/entrada/tabla_entrada.html')

def salidaDiseño(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Diseño y produccion")
                entradas = Salida.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'diseñoProduccion/salida/tabla_salida.html', data)
        else:
                return render(request, 'diseñoProduccion/salida/tabla_salida.html')                              

def oportunidadDiseño(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Diseño y produccion")
                entradas = Oportunidades.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'diseñoProduccion/oportunidad/tabla_oportunidad.html', data)
        else:
                return render(request, 'diseñoProduccion/oportunidad/tabla_oportunidad.html')      


#logistica


def EntradaLogistica(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Logistica")
                entradas = Entrada.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'logistica/entrada/tabla_entrada.html', data)
        else:
                return render(request, 'logistica/entrada/tabla_entrada.html')


def salidaLogistica(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Logistica")
                entradas = Salida.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'logistica/salida/tabla_salida.html', data)
        else:
                return render(request, 'logistica/salida/tabla_salida.html') 


def oportunidadLogistica(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Logistica")
                entradas = Oportunidades.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'logistica/oportunidad/tabla_oportunidad.html', data)
        else:
                return render(request, 'logistica/oportunidad/tabla_oportunidad.html')                                     


# compra


def entradaCompra(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Compra")
                entradas = Entrada.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'compra/entrada/tabla_entrada.html', data)
        else:
                return render(request, 'compra/entrada/tabla_entrada.html')      

def salidaCompra(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Compra")
                entradas = Salida.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'compra/salida/tabla_salida.html', data)
        else:
                return render(request, 'compra/salida/tabla_salida.html') 


def oportunidadesCompra(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Compra")
                entradas = Oportunidades.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'compra/oportunidad/tabla_oportunidad.html', data)
        else:
                return render(request, 'compra/salida/tabla_oportunidad.html') 

#Uso consumo

def entradaUsoConsumo(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Uso consumo")
                entradas = Entrada.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'usoConsumo/entrada/tabla_entrada.html', data)
        else:
                return render(request, 'usoConsumo/entrada/tabla_entrada.html') 


def salidaUsoConsumo(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Uso consumo")
                entradas = Salida.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'usoConsumo/salida/tabla_salida.html', data)
        else:
                return render(request, 'usoConsumo/salida/tabla_salida.html')


def oportunidadUsoConsumo(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Uso consumo")
                entradas = Oportunidades.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'usoConsumo/oportunidad/tabla_oportunidad.html', data)
        else:
                return render(request, 'usoConsumo/oportunidad/tabla_oportunidad.html')



#Fin de vida
def entradaFin(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Fin de vida")
                entradas = Entrada.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'finVida/entrada/tabla_entrada.html', data)
        else:
                return render(request, 'finVida/entrada/tabla_entrada.html')               

def salidaFin(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Fin de vida")
                entradas = Salida.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'finVida/salida/tabla_salida.html', data)
        else:
                return render(request, 'finVida/salida/tabla_salida.html')

def oportunidadFin(request):

        if request.user.is_authenticated:

                registros = RegistroTrabajador.objects.filter(usuario=request.user)
                etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(nombre="Fin de vida")
                entradas = Oportunidades.objects.all()
                empresaArea = RegistroTrabajador.objects.all()
                
                # etapa = Etapa.objects.values_list("id_etapa", flat=True).filter(activo=True)
                print(etapa)

                data = {

                'registros': registros,
                'entradas': entradas,
                'empresaArea':empresaArea

                }
       
                return render(request,'finVida/oportunidad/tabla_oportunidad.html', data)
        else:
                return render(request, 'finVida/oportunidad/tabla_oportunidad.html')






# ############################################# graficos ###############################################

def homeGraficos(request):
        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        

        data = {
                'registros': registros,
                'empresas': empresas 
                }

        return render(request, 'graficos/home_graficos.html', data)



def etapaGraficos(request, id):
        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        area = AreaEmpresa.objects.filter(id_empresa = id)
        empresa = Empresa.objects.filter(id_empresa = id)
        

        data = {
                'registros': registros,
                'empresas': empresas,
                'area':area,
                'empresa':empresa,
                }

        return render(request, 'graficos/etapas.html', data)



def areasExtraccion(request, id):
        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        empresa = Empresa.objects.filter(id_empresa = id)
        areas = AreaEmpresa.objects.filter(id_empresa = id)
           


        data = {
                'registros': registros,
                'empresas': empresas,
                'empresa':empresa,
                'areas':areas
                }

        return render(request, 'extraccion/grafico/areas.html', data)




def graficosExtraccion(request, id):

        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        area_grafico = AreaEmpresa.objects.filter(id_area = id)
     
           
        ################################ fechas ################################################
        #mosrar en el vertice x del graficp la fechas de ingreso de las notas(entrada, salida, oportunidad) sin repetir las fechas

        etapa = Etapa.objects.get(nombre = "Extraccion materia prima") #trar solo la ID de la etapa "Extraccion materia prima"
        entradas = Entrada.objects.filter(etapa_id = etapa, id_area = id)
        salidas = Salida.objects.filter(etapa_id = etapa, id_area = id)
        oportunidades = Oportunidades.objects.filter(etapa_id = etapa, id_area = id)

        
        b1 = 0
        for i in area_grafico :
                if b1 < 1 :
                        empresa_id = i.id_empresa_id
                        b1 = b1 + 1


        empresa = Empresa.objects.filter(id_empresa = empresa_id)
        areas = AreaEmpresa.objects.filter(id_empresa = empresa_id)


        dias_total = []
     
        for e in entradas:
                dias_total.append(e.fecha)

        for e in salidas:
                dias_total.append(e.fecha)

        for e in oportunidades:
                dias_total.append(e.fecha)                  
       

        #print("list original", dias_total)

        convert_list_to_set = set(dias_total)
        #print("Set is: ",convert_list_to_set)

        new_list = list(convert_list_to_set)
        #print("Resultant List is: ",new_list)

        dias_total = list(convert_list_to_set)
        #print("Removed duplicates from original list: ",dias_total)

        dias_total.sort()


        #crear diccionario para grafico
        diccionario = {}

        for i in dias_total:
                diccionario[i] = 0

        for d in diccionario:
                for e in entradas:
                        if d == e.fecha:
                                diccionario[d] = diccionario[d] +1
                                print(diccionario[d])

        print(diccionario)
        clave_dicc = diccionario.keys()
        valor_dicc = diccionario.values()
        cantidad_datos_dicc = diccionario.items()


        ######## entradas ##############
        fechas = []
        for e in entradas:
                fechas.append(e.fecha)

        c = collections.Counter(fechas)
           

        clave = c.keys()
        valor = c.values()
        cantidad_datos = c.items()    


        ######## salidas ##############

        dicc_salidas= {}

        for i in dias_total:
                dicc_salidas[i] = 0

        for d in dicc_salidas:
                for e in salidas:
                        if d == e.fecha:
                                dicc_salidas[d] = dicc_salidas[d] +1
                                

        clave_dicc_salida = dicc_salidas.keys()
        valor_dicc_salida = dicc_salidas.values()
        cantidad_datos_dicc_salida = dicc_salidas.items()


        ######## oportunidades ##############

        dicc_oportunidades= {}

        for i in dias_total: #en este for creo un diccionario donde la i sea la key y el value sea 0 para todos los registros
                dicc_oportunidades[i] = 0 

        for d in dicc_oportunidades:
                for e in oportunidades:
                        if d == e.fecha:
                                dicc_oportunidades[d] = dicc_oportunidades[d] +1
                                

        clave_dicc_oportunidad = dicc_oportunidades.keys()
        valor_dicc_oportunidad = dicc_oportunidades.values()
        cantidad_datos_dicc_oportunidad = dicc_oportunidades.items()
              

        
      
        data = {
                'registros': registros,
                'empresas': empresas,
                'area_grafico':area_grafico,
                'areas':areas,
                'empresa':empresa,
                'dias_total':dias_total,
                'clave': clave,
                'valor': valor,
                'cantidad_datos' : cantidad_datos,
                'clave_dicc': clave_dicc,
                'valor_dicc': valor_dicc,
                'cantidad_datos_dicc' : cantidad_datos_dicc,
                'clave_dicc_salida' : clave_dicc_salida,
                'valor_dicc_salida' : valor_dicc_salida,
                'cantidad_datos_dicc_salida' : cantidad_datos_dicc_salida,
                'clave_dicc_oportunidad' : clave_dicc_oportunidad,
                'valor_dicc_oportunidad' : valor_dicc_oportunidad,
                'cantidad_datos_dicc_oportunidad' : cantidad_datos_dicc_oportunidad,
                

                
                }

        return render(request, 'extraccion/grafico/graficos.html', data)       


def areasDiseño(request, id):
        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        empresa = Empresa.objects.filter(id_empresa = id)
        areas = AreaEmpresa.objects.filter(id_empresa = id)
           


        data = {
                'registros': registros,
                'empresas': empresas,
                'empresa':empresa,
                'areas':areas
                }

        return render(request, 'diseñoProduccion/grafico/areas.html', data)




def graficosDiseño(request, id):

        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        area_grafico = AreaEmpresa.objects.filter(id_area = id)
        
           
        ################################ fechas ################################################
        #mosrar en el vertice x del graficp la fechas de ingreso de las notas(entrada, salida, oportunidad) sin repetir las fechas

        etapa = Etapa.objects.get(nombre = "Diseño y produccion") #trar solo la ID de la etapa "Extraccion materia prima"
        entradas = Entrada.objects.filter(etapa_id = etapa, id_area = id)
        salidas = Salida.objects.filter(etapa_id = etapa, id_area = id)
        oportunidades = Oportunidades.objects.filter(etapa_id = etapa, id_area = id)

        
        b1 = 0
        for i in area_grafico :
                if b1 < 1 :
                        empresa_id = i.id_empresa_id
                        b1 = b1 + 1

        empresa = Empresa.objects.filter(id_empresa = empresa_id)
        areas = AreaEmpresa.objects.filter(id_empresa = empresa_id)
            


        dias_total = []
     
        for e in entradas:
                dias_total.append(e.fecha)

        for e in salidas:
                dias_total.append(e.fecha)

        for e in oportunidades:
                dias_total.append(e.fecha)                  
       

        #print("list original", dias_total)

        convert_list_to_set = set(dias_total)
        #print("Set is: ",convert_list_to_set)

        new_list = list(convert_list_to_set)
        #print("Resultant List is: ",new_list)

        dias_total = list(convert_list_to_set)
        #print("Removed duplicates from original list: ",dias_total)

        dias_total.sort()



        ######## entradas ##############
        #crear diccionario para grafico
        diccionario = {}

        for i in dias_total:
                diccionario[i] = 0

        for d in diccionario:
                for e in entradas:
                        if d == e.fecha:
                                diccionario[d] = diccionario[d] +1
                                print(diccionario[d])

        print(diccionario)
        clave_dicc = diccionario.keys()
        valor_dicc = diccionario.values()
        cantidad_datos_dicc = diccionario.items()



        ######## salidas ##############

        dicc_salidas= {}

        for i in dias_total:
                dicc_salidas[i] = 0

        for d in dicc_salidas:
                for e in salidas:
                        if d == e.fecha:
                                dicc_salidas[d] = dicc_salidas[d] +1
                                

        clave_dicc_salida = dicc_salidas.keys()
        valor_dicc_salida = dicc_salidas.values()
        cantidad_datos_dicc_salida = dicc_salidas.items()


        ######## oportunidades ##############

        dicc_oportunidades= {}

        for i in dias_total: #en este for creo un diccionario donde la i sea la key y el value sea 0 para todos los registros
                dicc_oportunidades[i] = 0 

        for d in dicc_oportunidades:
                for e in oportunidades:
                        if d == e.fecha:
                                dicc_oportunidades[d] = dicc_oportunidades[d] +1
                                

        clave_dicc_oportunidad = dicc_oportunidades.keys()
        valor_dicc_oportunidad = dicc_oportunidades.values()
        cantidad_datos_dicc_oportunidad = dicc_oportunidades.items()
              

        
      
        data = {
                'registros': registros,
                'empresas': empresas,
                'area_grafico':area_grafico,
                'areas':areas,
                'empresa':empresa,
                'dias_total':dias_total,
                'clave_dicc': clave_dicc,
                'valor_dicc': valor_dicc,
                'cantidad_datos_dicc' : cantidad_datos_dicc,
                'clave_dicc_salida' : clave_dicc_salida,
                'valor_dicc_salida' : valor_dicc_salida,
                'cantidad_datos_dicc_salida' : cantidad_datos_dicc_salida,
                'clave_dicc_oportunidad' : clave_dicc_oportunidad,
                'valor_dicc_oportunidad' : valor_dicc_oportunidad,
                'cantidad_datos_dicc_oportunidad' : cantidad_datos_dicc_oportunidad,
                

                
                }

        return render(request, 'diseñoProduccion/grafico/graficos.html', data)       




def areasLogistica(request, id):
        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        empresa = Empresa.objects.filter(id_empresa = id)
        areas = AreaEmpresa.objects.filter(id_empresa = id)
           


        data = {
                'registros': registros,
                'empresas': empresas,
                'empresa':empresa,
                'areas':areas
                }

        return render(request, 'logistica/grafico/areas.html', data)        




def graficosLogistica(request, id):

        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        area_grafico = AreaEmpresa.objects.filter(id_area = id)
        
           
        ################################ fechas ################################################
        #mosrar en el vertice x del graficp la fechas de ingreso de las notas(entrada, salida, oportunidad) sin repetir las fechas

        etapa = Etapa.objects.get(nombre = "Logistica") #trar solo la ID de la etapa "Extraccion materia prima"
        entradas = Entrada.objects.filter(etapa_id = etapa, id_area = id)
        salidas = Salida.objects.filter(etapa_id = etapa, id_area = id)
        oportunidades = Oportunidades.objects.filter(etapa_id = etapa, id_area = id)

        
        b1 = 0
        for i in area_grafico :
                if b1 < 1 :
                        empresa_id = i.id_empresa_id
                        b1 = b1 + 1

        empresa = Empresa.objects.filter(id_empresa = empresa_id)
        areas = AreaEmpresa.objects.filter(id_empresa = empresa_id)
            


        dias_total = []
     
        for e in entradas:
                dias_total.append(e.fecha)

        for e in salidas:
                dias_total.append(e.fecha)

        for e in oportunidades:
                dias_total.append(e.fecha)                  
       

        #print("list original", dias_total)

        convert_list_to_set = set(dias_total)
        #print("Set is: ",convert_list_to_set)

        new_list = list(convert_list_to_set)
        #print("Resultant List is: ",new_list)

        dias_total = list(convert_list_to_set)
        #print("Removed duplicates from original list: ",dias_total)

        dias_total.sort()



        ######## entradas ##############
        #crear diccionario para grafico
        diccionario = {}

        for i in dias_total:
                diccionario[i] = 0

        for d in diccionario:
                for e in entradas:
                        if d == e.fecha:
                                diccionario[d] = diccionario[d] +1
                                print(diccionario[d])

        print(diccionario)
        clave_dicc = diccionario.keys()
        valor_dicc = diccionario.values()
        cantidad_datos_dicc = diccionario.items()



        ######## salidas ##############

        dicc_salidas= {}

        for i in dias_total:
                dicc_salidas[i] = 0

        for d in dicc_salidas:
                for e in salidas:
                        if d == e.fecha:
                                dicc_salidas[d] = dicc_salidas[d] +1
                                

        clave_dicc_salida = dicc_salidas.keys()
        valor_dicc_salida = dicc_salidas.values()
        cantidad_datos_dicc_salida = dicc_salidas.items()


        ######## oportunidades ##############

        dicc_oportunidades= {}

        for i in dias_total: #en este for creo un diccionario donde la i sea la key y el value sea 0 para todos los registros
                dicc_oportunidades[i] = 0 

        for d in dicc_oportunidades:
                for e in oportunidades:
                        if d == e.fecha:
                                dicc_oportunidades[d] = dicc_oportunidades[d] +1
                                

        clave_dicc_oportunidad = dicc_oportunidades.keys()
        valor_dicc_oportunidad = dicc_oportunidades.values()
        cantidad_datos_dicc_oportunidad = dicc_oportunidades.items()
              

        
      
        data = {
                'registros': registros,
                'empresas': empresas,
                'area_grafico':area_grafico,
                'areas':areas,
                'empresa':empresa,
                'dias_total':dias_total,
                'clave_dicc': clave_dicc,
                'valor_dicc': valor_dicc,
                'cantidad_datos_dicc' : cantidad_datos_dicc,
                'clave_dicc_salida' : clave_dicc_salida,
                'valor_dicc_salida' : valor_dicc_salida,
                'cantidad_datos_dicc_salida' : cantidad_datos_dicc_salida,
                'clave_dicc_oportunidad' : clave_dicc_oportunidad,
                'valor_dicc_oportunidad' : valor_dicc_oportunidad,
                'cantidad_datos_dicc_oportunidad' : cantidad_datos_dicc_oportunidad,
                

                
                }

        return render(request, 'logistica/grafico/graficos.html', data)         



def areasCompra(request, id):
        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        empresa = Empresa.objects.filter(id_empresa = id)
        areas = AreaEmpresa.objects.filter(id_empresa = id)
           


        data = {
                'registros': registros,
                'empresas': empresas,
                'empresa':empresa,
                'areas':areas
                }

        return render(request, 'compra/grafico/areas.html', data)        




def graficosCompra(request, id):

        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        area_grafico = AreaEmpresa.objects.filter(id_area = id)
        
           
        ################################ fechas ################################################
        #mosrar en el vertice x del graficp la fechas de ingreso de las notas(entrada, salida, oportunidad) sin repetir las fechas

        etapa = Etapa.objects.get(nombre = "Compra") #trar solo la ID de la etapa "Extraccion materia prima"
        entradas = Entrada.objects.filter(etapa_id = etapa, id_area = id)
        salidas = Salida.objects.filter(etapa_id = etapa, id_area = id)
        oportunidades = Oportunidades.objects.filter(etapa_id = etapa, id_area = id)

        
        b1 = 0
        for i in area_grafico :
                if b1 < 1 :
                        empresa_id = i.id_empresa_id
                        b1 = b1 + 1

        empresa = Empresa.objects.filter(id_empresa = empresa_id)
        areas = AreaEmpresa.objects.filter(id_empresa = empresa_id)    


        dias_total = []
     
        for e in entradas:
                dias_total.append(e.fecha)

        for e in salidas:
                dias_total.append(e.fecha)

        for e in oportunidades:
                dias_total.append(e.fecha)                  
       

        #print("list original", dias_total)

        convert_list_to_set = set(dias_total)
        #print("Set is: ",convert_list_to_set)

        new_list = list(convert_list_to_set)
        #print("Resultant List is: ",new_list)

        dias_total = list(convert_list_to_set)
        #print("Removed duplicates from original list: ",dias_total)

        dias_total.sort()



        ######## entradas ##############
        #crear diccionario para grafico
        diccionario = {}

        for i in dias_total:
                diccionario[i] = 0

        for d in diccionario:
                for e in entradas:
                        if d == e.fecha:
                                diccionario[d] = diccionario[d] +1
                                print(diccionario[d])

        print(diccionario)
        clave_dicc = diccionario.keys()
        valor_dicc = diccionario.values()
        cantidad_datos_dicc = diccionario.items()



        ######## salidas ##############

        dicc_salidas= {}

        for i in dias_total:
                dicc_salidas[i] = 0

        for d in dicc_salidas:
                for e in salidas:
                        if d == e.fecha:
                                dicc_salidas[d] = dicc_salidas[d] +1
                                

        clave_dicc_salida = dicc_salidas.keys()
        valor_dicc_salida = dicc_salidas.values()
        cantidad_datos_dicc_salida = dicc_salidas.items()


        ######## oportunidades ##############

        dicc_oportunidades= {}

        for i in dias_total: #en este for creo un diccionario donde la i sea la key y el value sea 0 para todos los registros
                dicc_oportunidades[i] = 0 

        for d in dicc_oportunidades:
                for e in oportunidades:
                        if d == e.fecha:
                                dicc_oportunidades[d] = dicc_oportunidades[d] +1
                                

        clave_dicc_oportunidad = dicc_oportunidades.keys()
        valor_dicc_oportunidad = dicc_oportunidades.values()
        cantidad_datos_dicc_oportunidad = dicc_oportunidades.items()
              

        
      
        data = {
                'registros': registros,
                'empresas': empresas,
                'area_grafico':area_grafico,
                'areas':areas,
                'empresa':empresa,
                'dias_total':dias_total,
                'clave_dicc': clave_dicc,
                'valor_dicc': valor_dicc,
                'cantidad_datos_dicc' : cantidad_datos_dicc,
                'clave_dicc_salida' : clave_dicc_salida,
                'valor_dicc_salida' : valor_dicc_salida,
                'cantidad_datos_dicc_salida' : cantidad_datos_dicc_salida,
                'clave_dicc_oportunidad' : clave_dicc_oportunidad,
                'valor_dicc_oportunidad' : valor_dicc_oportunidad,
                'cantidad_datos_dicc_oportunidad' : cantidad_datos_dicc_oportunidad,
                

                
                }

        return render(request, 'compra/grafico/graficos.html', data)     



def areasUso(request, id):
        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        empresa = Empresa.objects.filter(id_empresa = id)
        areas = AreaEmpresa.objects.filter(id_empresa = id)
           


        data = {
                'registros': registros,
                'empresas': empresas,
                'empresa':empresa,
                'areas':areas
                }

        return render(request, 'usoConsumo/grafico/areas.html', data)        



def graficosUso(request, id):

        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        area_grafico = AreaEmpresa.objects.filter(id_area = id)
        
           
        ################################ fechas ################################################
        #mosrar en el vertice x del graficp la fechas de ingreso de las notas(entrada, salida, oportunidad) sin repetir las fechas

        etapa = Etapa.objects.get(nombre = "Uso consumo") #trar solo la ID de la etapa "Extraccion materia prima"
        entradas = Entrada.objects.filter(etapa_id = etapa, id_area = id)
        salidas = Salida.objects.filter(etapa_id = etapa, id_area = id)
        oportunidades = Oportunidades.objects.filter(etapa_id = etapa, id_area = id)

        
        b1 = 0
        for i in area_grafico :
                if b1 < 1 :
                        empresa_id = i.id_empresa_id
                        b1 = b1 + 1

        empresa = Empresa.objects.filter(id_empresa = empresa_id)
        areas = AreaEmpresa.objects.filter(id_empresa = empresa_id)  
            


        dias_total = []
     
        for e in entradas:
                dias_total.append(e.fecha)

        for e in salidas:
                dias_total.append(e.fecha)

        for e in oportunidades:
                dias_total.append(e.fecha)                  
       

        #print("list original", dias_total)

        convert_list_to_set = set(dias_total)
        #print("Set is: ",convert_list_to_set)

        new_list = list(convert_list_to_set)
        #print("Resultant List is: ",new_list)

        dias_total = list(convert_list_to_set)
        #print("Removed duplicates from original list: ",dias_total)

        dias_total.sort()



        ######## entradas ##############
        #crear diccionario para grafico
        diccionario = {}

        for i in dias_total:
                diccionario[i] = 0

        for d in diccionario:
                for e in entradas:
                        if d == e.fecha:
                                diccionario[d] = diccionario[d] +1
                                print(diccionario[d])

        print(diccionario)
        clave_dicc = diccionario.keys()
        valor_dicc = diccionario.values()
        cantidad_datos_dicc = diccionario.items()



        ######## salidas ##############

        dicc_salidas= {}

        for i in dias_total:
                dicc_salidas[i] = 0

        for d in dicc_salidas:
                for e in salidas:
                        if d == e.fecha:
                                dicc_salidas[d] = dicc_salidas[d] +1
                                

        clave_dicc_salida = dicc_salidas.keys()
        valor_dicc_salida = dicc_salidas.values()
        cantidad_datos_dicc_salida = dicc_salidas.items()


        ######## oportunidades ##############

        dicc_oportunidades= {}

        for i in dias_total: #en este for creo un diccionario donde la i sea la key y el value sea 0 para todos los registros
                dicc_oportunidades[i] = 0 

        for d in dicc_oportunidades:
                for e in oportunidades:
                        if d == e.fecha:
                                dicc_oportunidades[d] = dicc_oportunidades[d] +1
                                

        clave_dicc_oportunidad = dicc_oportunidades.keys()
        valor_dicc_oportunidad = dicc_oportunidades.values()
        cantidad_datos_dicc_oportunidad = dicc_oportunidades.items()
              

        
      
        data = {
                'registros': registros,
                'empresas': empresas,
                'area_grafico':area_grafico,
                'areas':areas,
                'empresa':empresa,
                'dias_total':dias_total,
                'clave_dicc': clave_dicc,
                'valor_dicc': valor_dicc,
                'cantidad_datos_dicc' : cantidad_datos_dicc,
                'clave_dicc_salida' : clave_dicc_salida,
                'valor_dicc_salida' : valor_dicc_salida,
                'cantidad_datos_dicc_salida' : cantidad_datos_dicc_salida,
                'clave_dicc_oportunidad' : clave_dicc_oportunidad,
                'valor_dicc_oportunidad' : valor_dicc_oportunidad,
                'cantidad_datos_dicc_oportunidad' : cantidad_datos_dicc_oportunidad,
                

                
                }

        return render(request, 'usoConsumo/grafico/graficos.html', data)     




def areasFin(request, id):
        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        empresa = Empresa.objects.filter(id_empresa = id)
        areas = AreaEmpresa.objects.filter(id_empresa = id)
           


        data = {
                'registros': registros,
                'empresas': empresas,
                'empresa':empresa,
                'areas':areas
                }

        return render(request, 'finVida/grafico/areas.html', data)      


def graficosFin(request, id):

        registros = RegistroTrabajador.objects.filter(usuario = request.user)
        empresas = Empresa.objects.all()
        area_grafico = AreaEmpresa.objects.filter(id_area = id)
        
           
        ################################ fechas ################################################
        #mosrar en el vertice x del graficp la fechas de ingreso de las notas(entrada, salida, oportunidad) sin repetir las fechas

        etapa = Etapa.objects.get(nombre = "Fin de vida") #trar solo la ID de la etapa "Extraccion materia prima"
        entradas = Entrada.objects.filter(etapa_id = etapa, id_area = id)
        salidas = Salida.objects.filter(etapa_id = etapa, id_area = id)
        oportunidades = Oportunidades.objects.filter(etapa_id = etapa, id_area = id)

        
        b1 = 0
        for i in area_grafico :
                if b1 < 1 :
                        empresa_id = i.id_empresa_id
                        b1 = b1 + 1

        empresa = Empresa.objects.filter(id_empresa = empresa_id)
        areas = AreaEmpresa.objects.filter(id_empresa = empresa_id)  
            


        dias_total = []
     
        for e in entradas:
                dias_total.append(e.fecha)

        for e in salidas:
                dias_total.append(e.fecha)

        for e in oportunidades:
                dias_total.append(e.fecha)                  
       

        #print("list original", dias_total)

        convert_list_to_set = set(dias_total)
        #print("Set is: ",convert_list_to_set)

        new_list = list(convert_list_to_set)
        #print("Resultant List is: ",new_list)

        dias_total = list(convert_list_to_set)
        #print("Removed duplicates from original list: ",dias_total)

        dias_total.sort()



        ######## entradas ##############
        #crear diccionario para grafico
        diccionario = {}

        for i in dias_total:
                diccionario[i] = 0

        for d in diccionario:
                for e in entradas:
                        if d == e.fecha:
                                diccionario[d] = diccionario[d] +1
                                print(diccionario[d])

        print(diccionario)
        clave_dicc = diccionario.keys()
        valor_dicc = diccionario.values()
        cantidad_datos_dicc = diccionario.items()



        ######## salidas ##############

        dicc_salidas= {}

        for i in dias_total:
                dicc_salidas[i] = 0

        for d in dicc_salidas:
                for e in salidas:
                        if d == e.fecha:
                                dicc_salidas[d] = dicc_salidas[d] +1
                                

        clave_dicc_salida = dicc_salidas.keys()
        valor_dicc_salida = dicc_salidas.values()
        cantidad_datos_dicc_salida = dicc_salidas.items()


        ######## oportunidades ##############

        dicc_oportunidades= {}

        for i in dias_total: #en este for creo un diccionario donde la i sea la key y el value sea 0 para todos los registros
                dicc_oportunidades[i] = 0 

        for d in dicc_oportunidades:
                for e in oportunidades:
                        if d == e.fecha:
                                dicc_oportunidades[d] = dicc_oportunidades[d] +1
                                

        clave_dicc_oportunidad = dicc_oportunidades.keys()
        valor_dicc_oportunidad = dicc_oportunidades.values()
        cantidad_datos_dicc_oportunidad = dicc_oportunidades.items()
              

        
      
        data = {
                'registros': registros,
                'empresas': empresas,
                'area_grafico':area_grafico,
                'areas':areas,
                'empresa':empresa,
                'dias_total':dias_total,
                'clave_dicc': clave_dicc,
                'valor_dicc': valor_dicc,
                'cantidad_datos_dicc' : cantidad_datos_dicc,
                'clave_dicc_salida' : clave_dicc_salida,
                'valor_dicc_salida' : valor_dicc_salida,
                'cantidad_datos_dicc_salida' : cantidad_datos_dicc_salida,
                'clave_dicc_oportunidad' : clave_dicc_oportunidad,
                'valor_dicc_oportunidad' : valor_dicc_oportunidad,
                'cantidad_datos_dicc_oportunidad' : cantidad_datos_dicc_oportunidad,
                

                
                }

        return render(request, 'finVida/grafico/graficos.html', data)     





# reportes de excel

#  ///////////////////////////////////////// reportes de excel extraccion ///////////////////////////////////////////////////////////////////////////

class ReporteExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Extraccion materia prima") #trar solo la ID de la etapa "Extraccion materia prima"
        entradas = Entrada.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Entradas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 

        for e in entradas:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelSalida(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Extraccion materia prima") #trar solo la ID de la etapa "Extraccion materia prima"
        salidas = Salida.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Salidas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 #fila en la que comienza la tabla 

        for e in salidas:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelOportunidades(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Extraccion materia prima") #trar solo la ID de la etapa "Extraccion materia prima"
        oportunidades = Oportunidades.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Salidas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 #fila en la que comienza la tabla 

        for e in oportunidades:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response




#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

#  ///////////////////////////////////////// reportes de excel Diseño y produccion ///////////////////////////////////////////////////////////////////////////

class ReporteExcelEntradaDiseño(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Diseño y produccion") #trar solo la ID de la etapa "Extraccion materia prima"
        entradas = Entrada.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Entradas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4

        for e in entradas:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelSalidaDiseño(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Diseño y produccion") #trar solo la ID de la etapa "Extraccion materia prima"
        salidas = Salida.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Salidas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 #fila en la que comienza la tabla 

        for e in salidas:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelOportunidadDiseño(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Diseño y produccion") #trar solo la ID de la etapa "Extraccion materia prima"
        oportunidades = Oportunidades.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Salidas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 #fila en la que comienza la tabla 

        for e in oportunidades:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response




#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////


#  ///////////////////////////////////////// reportes de excel Logistica ///////////////////////////////////////////////////////////////////////////

class ReporteExcelEntradaLogistica(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Logistica") #trar solo la ID de la etapa "Extraccion materia prima"
        entradas = Entrada.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Entradas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 

        for e in entradas:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelSalidaLogistica(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Logistica") #trar solo la ID de la etapa "Extraccion materia prima"
        salidas = Salida.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Salidas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 #fila en la que comienza la tabla 

        for e in salidas:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelOportunidadLogistica(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Logistica") #trar solo la ID de la etapa "Extraccion materia prima"
        oportunidades = Oportunidades.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Salidas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 #fila en la que comienza la tabla 

        for e in oportunidades:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response




#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////



#  ///////////////////////////////////////// reportes de excel Compra ///////////////////////////////////////////////////////////////////////////

class ReporteExcelEntradaCompra(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Compra") #trar solo la ID de la etapa "Extraccion materia prima"
        entradas = Entrada.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Entradas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4

        for e in entradas:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelSalidaCompra(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Compra") #trar solo la ID de la etapa "Extraccion materia prima"
        salidas = Salida.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Salidas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 #fila en la que comienza la tabla 

        for e in salidas:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelOportunidadCompra(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Compra") #trar solo la ID de la etapa "Extraccion materia prima"
        oportunidades = Oportunidades.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Oportunidades'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 #fila en la que comienza la tabla 

        for e in oportunidades:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response




#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////


#  ///////////////////////////////////////// reportes de excel Uso Consumo ///////////////////////////////////////////////////////////////////////////

class ReporteExcelEntradaUso(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Uso consumo") #trar solo la ID de la etapa "Extraccion materia prima"
        entradas = Entrada.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Entradas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4

        for e in entradas:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelSalidaUso(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Uso consumo") #trar solo la ID de la etapa "Extraccion materia prima"
        salidas = Salida.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Salidas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 #fila en la que comienza la tabla 

        for e in salidas:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelOportunidadUso(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Uso consumo") #trar solo la ID de la etapa "Extraccion materia prima"
        oportunidades = Oportunidades.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Oportunidades'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 #fila en la que comienza la tabla 

        for e in oportunidades:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////


#  ///////////////////////////////////////// reportes de excel Fin de vida  ///////////////////////////////////////////////////////////////////////////

class ReporteExcelEntradaFin(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Fin de vida") #trar solo la ID de la etapa "Extraccion materia prima"
        entradas = Entrada.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Entradas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4

        for e in entradas:
                print(e.id_area.id_empresa_id)
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelSalidaFin(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Fin de vida") #trar solo la ID de la etapa "Extraccion materia prima"
        salidas = Salida.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Salidas'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 #fila en la que comienza la tabla 

        for e in salidas:
                
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelOportunidadFin(TemplateView):
    def get(self, request, *args, **kwargs):
        etapa = Etapa.objects.get(nombre = "Fin de vida") #trar solo la ID de la etapa "Extraccion materia prima"
        oportunidades = Oportunidades.objects.filter(etapa_id = etapa)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Oportunidades'

        ws.merge_cells('B1:G1')
        ws['B3'] = 'ID Empresa'
        ws['C3'] = 'ID Area'
        ws['D3'] = 'ID Usuario'
        ws['E3'] = 'Nombre'
        ws['F3'] = 'Fecha'
        ws['G3'] = 'ID Etapa'

        cont = 4 #fila en la que comienza la tabla 

        for e in oportunidades:
                ws.cell(row = cont, column = 2).value = e.id_area.id_empresa_id
                ws.cell(row = cont, column = 3).value = e.id_area_id
                ws.cell(row = cont, column = 4).value = e.usuario_id
                ws.cell(row = cont, column = 5).value = e.nombre
                ws.cell(row = cont, column = 6).value = e.fecha
                ws.cell(row = cont, column = 7).value = e.etapa_id
                cont+=1

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response



def log_telegan(request):
        registros = RegistroTrabajador.objects.filter(usuario=request.user)
        logs_telegram = LogTelegram.objects.all()

        data = {
                'logs_telegram':logs_telegram,
                'registros': registros
        }
        return render(request,'log_telegram/log_telegram.html', data)



